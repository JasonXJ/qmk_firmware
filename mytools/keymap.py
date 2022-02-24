#!/usr/bin/env python3
from pathlib import Path
import re
from dataclasses import dataclass

import openpyxl
import click


DIR = Path(__file__).resolve().parent
ROOT_DIR = DIR.parent
KEYMAP_NAME = 'jasonxj'

SHEET_ROWS_PER_LAYER = 8
SHEET_COLUMNS = 14
CELL_VALUE_MAP = {
    '-': '_______',
    'x': 'XXXXXXX',
}

# NA means the cell should contains the name of the layer.
# 
# - OR means the key index is in ORder. It will count from 0, and skip all the
#   indices that were explicitly specified. These should be replaced by actual
#   index after being handled by `process_pattern()`.
# - EM means there should be no keys there (i.e. EMpty).
# - IG means IGnore
# - Otherwise, it should be the index of the key.
NA = object()
OR = object()
EM = object()
IG = object()


@dataclass(frozen=True)
class Info:
    key_count: int
    layout_function: str
    pattern: [[object]]

    def __init__(self, key_count, layout_function, raw_pattern):
        object.__setattr__(self, 'key_count', key_count)
        object.__setattr__(self, 'layout_function', layout_function)
        object.__setattr__(self, 'pattern', self.process_pattern(raw_pattern))
        self._check()

    def _check(self):
        assert len(self.pattern) == SHEET_ROWS_PER_LAYER
        assert all(len(r) == SHEET_COLUMNS for r in self.pattern)
        keys = 0
        for row in self.pattern:
            for x in row:
                if x not in (EM, NA, IG):
                    keys += 1
        assert keys == self.key_count

    @classmethod
    def process_pattern(cls, raw_pattern):
        """Process a raw_pattern. After the processing, all the 'OR' should be replaced with an actual number"""
        occupied_numbers = set()
        for row in raw_pattern:
            for item in row:
                if isinstance(item, int):
                    occupied_numbers.add(item)
        pattern = []
        next_number = 0
        for row in raw_pattern:
            new_row = []
            pattern.append(new_row)
            for item in row:
                if item is OR:
                    while next_number in occupied_numbers: 
                        next_number += 1
                    new_row.append(next_number)
                    next_number += 1
                else:
                    new_row.append(item)
        return pattern


INFO_MAP = {
    'ergodox_ez': Info(76, 'LAYOUT_ergodox_pretty', [
        [NA, EM, EM, EM, EM, EM, EM, EM, EM, EM, EM, EM, EM, EM],  # Title (layer name)
        [OR, OR, OR, OR, OR, OR, OR, OR, OR, OR, OR, OR, OR, OR],
        [OR, OR, OR, OR, OR, OR, OR, OR, OR, OR, OR, OR, OR, OR],
        [OR, OR, OR, OR, OR, OR, EM, EM, OR, OR, OR, OR, OR, OR],
        [OR, OR, OR, OR, OR, OR, OR, OR, OR, OR, OR, OR, OR, OR],
        [OR, OR, OR, OR, OR, 64, 65, 66, 67, OR, OR, OR, OR, OR],
        [EM, EM, EM, EM, 70, 71, 68, 69, 74, 75, EM, EM, EM, EM],
        [EM, EM, EM, EM, EM, EM, 72, 73, EM, EM, EM, EM, EM, EM],
    ]),
    'moonlander': Info(72, 'LAYOUT_moonlander', [
        [NA, EM, EM, EM, EM, EM, EM, EM, EM, EM, EM, EM, EM, EM],  # Title (layer name)
        [OR, OR, OR, OR, OR, OR, OR, OR, OR, OR, OR, OR, OR, OR],
        [OR, OR, OR, OR, OR, OR, OR, OR, OR, OR, OR, OR, OR, OR],
        [OR, OR, OR, OR, OR, OR, EM, EM, OR, OR, OR, OR, OR, OR],
        [OR, OR, OR, OR, OR, OR, 34, 35, OR, OR, OR, OR, OR, OR],
        [OR, OR, OR, OR, OR, OR, IG, IG, OR, OR, OR, OR, OR, OR],
        [EM, EM, EM, EM, OR, OR, IG, IG, OR, OR, EM, EM, EM, EM],
        [EM, EM, EM, EM, EM, EM, 68, 69, EM, EM, EM, EM, EM, EM],
    ]),
}

@dataclass
class GenCode:
    keyboard_name: str
    keymap: str = ''
    config: str = ''

    def write(self):
        # TODO: d
        dir = ROOT_DIR / 'keyboards' / self.keyboard_name / 'keymaps' / KEYMAP_NAME
        with (dir / 'keymap-gen.c').open('w') as f:
            f.write(self.keymap)
        with (dir / 'config-gen.h').open('w') as f:
            f.write(self.config)


@dataclass
class Layer:
    name: str
    keys: list[str]
    

# TODO: check layer names?
@dataclass
class Keymap:
    layers: list[Layer]
    info: Info

    def append_code(self, gen_code: GenCode):
        gen_code.keymap += f"{self._c_layers_enum()}\n\n{self._c_keymaps()}\n\n"

    def _c_layers_enum(self):
        s = 'enum layers {\n';
        for layer in self.layers:
            s += f'    {layer.name},\n';
        s += '};\n'
        
        return s

    def _c_keymaps(self):
        s = 'const uint16_t PROGMEM keymaps[][MATRIX_ROWS][MATRIX_COLS] = {\n'
        for layer in self.layers:
            s += f'  [{layer.name}] = {self.info.layout_function}(\n';
            # TODO: pretty print!
            s += f'    {layer.keys[0]}'
            for key in layer.keys[1:]:
                s += f', {key}'
            s += f'\n  ),\n'
        s += '};\n'

        return s


def convert_raw_key_value(value):
    value = value.strip()
    return CELL_VALUE_MAP.get(value, value)


def to_sheet_location(row, col):
    return chr(ord('A') + col) + str(row + 1)


def parse(filename: str, info: Info):
    wb = openpyxl.load_workbook(filename)
    ws = wb.active
    ws.calculate_dimension()

    def get_value(row, col):
        return ws.cell(row + 1, col + 1).value

    assert ws.max_column == SHEET_COLUMNS
    assert ws.max_row % SHEET_ROWS_PER_LAYER == 0

    layer_count = ws.max_row // SHEET_ROWS_PER_LAYER
    print(f"layer_count = {layer_count}")

    layers = []

    for layer_i in range(layer_count):
        start_row = SHEET_ROWS_PER_LAYER * layer_i
        name = None
        keys = [None] * info.key_count
        key_cursor = 0

        def set_key(index, value):
            assert keys[index] is None, f'keys[index] is not None. index={index}, key={keys[index]}'
            keys[index] = convert_raw_key_value(value)
        
        for row_offset in range(SHEET_ROWS_PER_LAYER):
            for col in range(SHEET_COLUMNS):
                row = start_row + row_offset

                def debug_cell_location():
                    return f"Cell {to_sheet_location(row, col)} ({row}, {col})"

                value = get_value(row, col)
                pattern_item = info.pattern[row_offset][col]

                if pattern_item is IG:
                    continue

                if pattern_item is EM:
                    if value is not None:
                        raise ValueError(f"{debug_cell_location()} has value {repr(value)}, but it should be empty")
                    continue

                if value is None:
                    raise ValueError(f"{debug_cell_location()} should have value")

                if pattern_item is NA:
                    assert name is None
                    name = value.strip()
                    continue

                assert isinstance(pattern_item, int)
                set_key(pattern_item, value)

        assert name is not None
        assert all(isinstance(k, str) and k for k in keys)
        layers.append(Layer(name, keys))

    return Keymap(layers, info)


@dataclass
class Combo:
    target: str
    source: list[str]
    name: str = None


COMBOS = [
    Combo('KC_TAB'   , ['KC_S', 'KC_D']),
    Combo('KC_ESC'   , ['KC_D', 'KC_F']),
    Combo('KC_BSPACE', ['KC_J', 'KC_K']),
    Combo('KC_ENTER' , ['KC_K', 'KC_L']),
    Combo('KC_MINUS' , ['KC_W', 'KC_E']),
    Combo('LSFT(KC_MINUS)' , ['KC_E', 'KC_R'], name='CB_SHIFT_MINUS'),
    # Combo('KC_EQUAL' , ['KC_I', 'KC_O']),
    # Combo('LSFT(KC_EQUAL)' , ['KC_U', 'KC_I'], name='CB_SHIFT_EQUAL'),
    Combo('KC_LBRACKET', ['KC_F', 'KC_G']),
    Combo('LSFT(KC_LBRACKET)', ['KC_V', 'KC_B'], name='CB_SHIFT_LBRACKET'),
    Combo('KC_RBRACKET', ['KC_H', 'KC_J']),
    Combo('LSFT(KC_RBRACKET)', ['KC_N', 'KC_M'], name='CB_SHIFT_RBRACKET'),
    Combo('KC_BSLASH', ['KC_M', 'KC_COMMA']),
]

_IDENTIFIER_PATTERN = re.compile(r'[A-Za-z_]\w*')

def is_identifier(s):
    return _IDENTIFIER_PATTERN.fullmatch(s) is not None


def append_combos_code(gen_code: GenCode, combos: list[Combo]):
    enum_def = 'enum combos {\n'
    progmen = ''
    array = 'combo_t key_combos[COMBO_COUNT] = {\n'
    names = set()
    sorted_sources = set()

    def insert_and_check_unique(s, value):
        assert value not in s
        s.add(value)

    for combo in combos:
        name = combo.name
        if name is None:
            name = f'CB_{combo.target}'
        assert is_identifier(name)
        insert_and_check_unique(names, name)
        insert_and_check_unique(sorted_sources, tuple(sorted(combo.source)))

        enum_def += f'    {name},\n'
        progmem_name = f'{name}_progmem'
        progmen += f'const uint16_t PROGMEM {progmem_name}[] = {{{", ".join(combo.source)}, COMBO_END}};\n'
        array += f'    [{name}] = COMBO({progmem_name}, {combo.target}),\n'

    enum_def += '};\n'
    array += '};\n'
    
    gen_code.config += f'#define COMBO_COUNT {len(combos)}\n\n'
    gen_code.keymap += f'{enum_def}{progmen}{array}'
        

@click.command()
@click.argument('keyboard_name')
def cli(keyboard_name):
    gen_code = GenCode(keyboard_name)

    keymap = parse(DIR / 'keymap.xlsx', INFO_MAP[keyboard_name])
    keymap.append_code(gen_code)

    append_combos_code(gen_code, COMBOS)

    gen_code.write()

    print(80*'-')
    print(gen_code.config)
    print(80*'-')
    print(gen_code.keymap)


if __name__ == "__main__":
    cli()
