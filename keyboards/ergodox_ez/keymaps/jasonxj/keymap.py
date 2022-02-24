#!/usr/bin/env python3
import re
from dataclasses import dataclass

import openpyxl


@dataclass
class GenCode:
    keymap: str = ''
    config: str = ''

    def write(self):
        with open('./keymap-gen.c', 'w') as f:
            f.write(self.keymap)
        with open('./config-gen.h', 'w') as f:
            f.write(self.config)


KEYS_COUNT = 76
SHEET_ROWS_PER_LAYER = 8
SHEET_COLUMNS = 14
CELL_VALUE_MAP = {
    '-': '_______',
    'x': 'XXXXXXX',
}

# NA means the cell should contains the name of the layer. OR means the key
# index is in ORder (counting from 0), EM means there should be no keys there
# (i.e. EMpty). Otherwise, it should be the index of the key.
NA = object()
OR = object()
EM = object()
SHEET_LAYER_PATTERN = [
    [NA, EM, EM, EM, EM, EM, EM, EM, EM, EM, EM, EM, EM, EM],  # Title (layer name)
    [OR, OR, OR, OR, OR, OR, OR, OR, OR, OR, OR, OR, OR, OR],
    [OR, OR, OR, OR, OR, OR, OR, OR, OR, OR, OR, OR, OR, OR],
    [OR, OR, OR, OR, OR, OR, EM, EM, OR, OR, OR, OR, OR, OR],
    [OR, OR, OR, OR, OR, OR, OR, OR, OR, OR, OR, OR, OR, OR],
    [OR, OR, OR, OR, OR, 64, 65, 66, 67, OR, OR, OR, OR, OR],
    [EM, EM, EM, EM, 70, 71, 68, 69, 74, 75, EM, EM, EM, EM],
    [EM, EM, EM, EM, EM, EM, 72, 73, EM, EM, EM, EM, EM, EM],
]
def _debug_check_pattern():
    assert len(SHEET_LAYER_PATTERN) == SHEET_ROWS_PER_LAYER
    assert all(len(r) == SHEET_COLUMNS for r in SHEET_LAYER_PATTERN)
    keys = 0
    for row in SHEET_LAYER_PATTERN:
        for x in row:
            if x is not EM and x is not NA:
                keys += 1
    assert keys == KEYS_COUNT
_debug_check_pattern()


@dataclass
class Layer:
    name: str
    keys: list[str]
    

# TODO: check layer names?
@dataclass
class Keymap:
    layers: list[Layer]

    def append_code(self, gen_code: GenCode):
        gen_code.keymap += f"{self._c_layers_enum()}\n\n{self._c_keymaps()}\n\n"

    def _c_layers_enum(self):
        s = 'enum layers {\n';
        for layer in self.layers:
            s += f'    {layer.name},\n';
        s += '};\n'
        
        return s

    def _c_keymaps(self):
        s = 'const uint16_t PROGMEM keymaps[][MATRIX_ROWS][MATRIX_COLS] = {\n'
        for layer in self.layers:
            s += f'  [{layer.name}] = LAYOUT_ergodox_pretty(\n';
            # TODO: pretty print!
            s += f'    {layer.keys[0]}'
            for key in layer.keys[1:]:
                s += f', {key}'
            s += f'\n  ),\n'
        s += '};\n'

        return s


def convert_raw_key_value(value):
    value = value.strip()
    return CELL_VALUE_MAP.get(value, value)


def to_sheet_location(row, col):
    return chr(ord('A') + col) + str(row + 1)


def parse(filename):
    wb = openpyxl.load_workbook(filename)
    ws = wb.active
    ws.calculate_dimension()

    def get_value(row, col):
        return ws.cell(row + 1, col + 1).value

    assert ws.max_column == SHEET_COLUMNS
    assert ws.max_row % SHEET_ROWS_PER_LAYER == 0

    layer_count = ws.max_row // SHEET_ROWS_PER_LAYER
    print(f"layer_count = {layer_count}")

    layers = []

    for layer_i in range(layer_count):
        start_row = SHEET_ROWS_PER_LAYER * layer_i
        name = None
        keys = [None] * KEYS_COUNT
        key_cursor = 0

        def set_key(index, value):
            assert keys[index] is None
            keys[index] = convert_raw_key_value(value)
        
        for row_offset in range(SHEET_ROWS_PER_LAYER):
            for col in range(SHEET_COLUMNS):
                row = start_row + row_offset

                def debug_cell_location():
                    return f"Cell {to_sheet_location(row, col)} ({row}, {col})"

                value = get_value(row, col)
                pattern = SHEET_LAYER_PATTERN[row_offset][col]
                if pattern is EM:
                    if value is not None:
                        raise ValueError(f"{debug_cell_location()} has value {repr(value)}, but it should be empty")
                    continue

                if value is None:
                    raise ValueError(f"{debug_cell_location()} should have value")

                if pattern is NA:
                    assert name is None
                    name = value.strip()
                    continue

                if pattern is OR:
                    set_key(key_cursor, value)
                    key_cursor += 1
                    continue

                assert isinstance(pattern, int)
                set_key(pattern, value)

        assert name is not None
        assert all(isinstance(k, str) and k for k in keys)
        layers.append(Layer(name, keys))

    return Keymap(layers)


@dataclass
class Combo:
    target: str
    source: list[str]
    name: str = None


COMBOS = [
    Combo('KC_TAB'   , ['KC_S', 'KC_D']),
    Combo('KC_ESC'   , ['KC_D', 'KC_F']),
    Combo('KC_BSPACE', ['KC_J', 'KC_K']),
    Combo('KC_ENTER' , ['KC_K', 'KC_L']),
    Combo('KC_MINUS' , ['KC_W', 'KC_E']),
    Combo('LSFT(KC_MINUS)' , ['KC_E', 'KC_R'], name='CB_SHIFT_MINUS'),
    # Combo('KC_EQUAL' , ['KC_I', 'KC_O']),
    # Combo('LSFT(KC_EQUAL)' , ['KC_U', 'KC_I'], name='CB_SHIFT_EQUAL'),
    Combo('KC_LBRACKET', ['KC_F', 'KC_G']),
    Combo('LSFT(KC_LBRACKET)', ['KC_V', 'KC_B'], name='CB_SHIFT_LBRACKET'),
    Combo('KC_RBRACKET', ['KC_H', 'KC_J']),
    Combo('LSFT(KC_RBRACKET)', ['KC_N', 'KC_M'], name='CB_SHIFT_RBRACKET'),
    Combo('KC_BSLASH', ['KC_M', 'KC_COMMA']),
]

_IDENTIFIER_PATTERN = re.compile(r'[A-Za-z_]\w*')

def is_identifier(s):
    return _IDENTIFIER_PATTERN.fullmatch(s) is not None


def append_combos_code(gen_code: GenCode, combos: list[Combo]):
    enum_def = 'enum combos {\n'
    progmen = ''
    array = 'combo_t key_combos[COMBO_COUNT] = {\n'
    names = set()
    sorted_sources = set()

    def insert_and_check_unique(s, value):
        assert value not in s
        s.add(value)

    for combo in combos:
        name = combo.name
        if name is None:
            name = f'CB_{combo.target}'
        assert is_identifier(name)
        insert_and_check_unique(names, name)
        insert_and_check_unique(sorted_sources, tuple(sorted(combo.source)))

        enum_def += f'    {name},\n'
        progmem_name = f'{name}_progmem'
        progmen += f'const uint16_t PROGMEM {progmem_name}[] = {{{", ".join(combo.source)}, COMBO_END}};\n'
        array += f'    [{name}] = COMBO({progmem_name}, {combo.target}),\n'

    enum_def += '};\n'
    array += '};\n'
    
    gen_code.config += f'#define COMBO_COUNT {len(combos)}\n\n'
    gen_code.keymap += f'{enum_def}{progmen}{array}'
        


if __name__ == "__main__":
    gen_code = GenCode()

    keymap = parse('./keymap.xlsx')
    keymap.append_code(gen_code)

    append_combos_code(gen_code, COMBOS)

    gen_code.write()

    print(80*'-')
    print(gen_code.config)
    print(80*'-')
    print(gen_code.keymap)
